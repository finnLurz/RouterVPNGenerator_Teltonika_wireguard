#!/usr/bin/env python3
"""
excel_to_yaml.py - Pflege per Excel, Erzeugung der standorte.yaml mit Validierung.

Ablauf:
  1. standorte.xlsx pflegen (Blatt 'Standorte', 'Anker', 'Zentrale_Netze')
  2. python3 excel_to_yaml.py        -> prueft + schreibt standorte.yaml
  3. python3 generate.py             -> Geraete-/Hub-Configs

Validierung (bricht bei Fehler ab, BEVOR etwas ausgerollt wird):
  - doppelte Standortnamen
  - doppelte site_index
  - doppelte oder ueberlappende LAN-Subnetze
  - ungueltiges Subnetz / Profil
  - site_index ausserhalb 1..254
"""

import ipaddress
import sys

try:
    import yaml
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Fehlt: pip install pyyaml openpyxl")

XLSX = "standorte.xlsx"
YAML = "standorte.yaml"
VALID_PROFILES = {"sirene", "fahrzeug"}


def read_rows(ws, headers):
    """Liest Zeilen ab der ersten Datenzeile, ignoriert Hinweiszeile/Leerzeilen."""
    out = []
    head_row = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    idx = {h: head_row.index(h) for h in headers if h in head_row}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        first = str(row[idx[headers[0]]]).strip() if row[idx[headers[0]]] else ""
        # Hinweiszeile (kursiv) ueberspringen: erkennt man am Beispieltext
        if first.lower() in ("eindeutiger name", "id"):
            continue
        rec = {}
        for h in headers:
            if h in idx and idx[h] < len(row):
                rec[h] = row[idx[h]]
        out.append(rec)
    return out


def fail(msg):
    print(f"FEHLER: {msg}")
    sys.exit(1)


def main():
    try:
        wb = load_workbook(XLSX, data_only=True)
    except FileNotFoundError:
        fail(f"{XLSX} nicht gefunden.")

    # --- bestehende YAML als Basis (global/profiles uebernehmen) ---
    try:
        base = yaml.safe_load(open(YAML, encoding="utf-8"))
    except FileNotFoundError:
        base = {"global": {"keepalive": 25, "tunnel_net_prefix": "10.80",
                           "hub_host_octet": 254, "watchdog_interval_s": 30,
                           "watchdog_fail_threshold": 3, "watchdog_recover_threshold": 4,
                           "routes_central": []},
                "profiles": {"sirene": {"lan_mask": 28, "fw_zone_lan": "lan"},
                             "fahrzeug": {"lan_mask": 27, "fw_zone_lan": "lan",
                                          "guest_subnet": "192.168.99.0/24"}}}

    # --- Standorte ---
    ws = wb["Standorte"]
    rows = read_rows(ws, ["name", "profile", "site_index", "lan_subnet", "aktiv"])
    # Auto-Vergabe: site_index (ab 11 Sirenen / ab 31 Fahrzeuge) und lan_subnet,
    # falls in der Tabelle leer gelassen. Du musst sie also NICHT pflegen.
    import ipaddress as _ip
    SIR_BASE=_ip.ip_network("192.168.80.0/24"); FZG_BASE=_ip.ip_network("192.168.81.0/24")
    sir_slots=list(SIR_BASE.subnets(new_prefix=28)); fzg_slots=list(FZG_BASE.subnets(new_prefix=27))
    sir_n=fzg_n=0; sir_idx=11; fzg_idx=31
    def _auto(rec):
        nonlocal_marker=None
    sites, names, indexes, nets = [], set(), set(), []
    for r in rows:
        name = str(r.get("name", "")).strip()
        if not name:
            continue
        prof = str(r.get("profile", "")).strip().lower()
        if prof not in VALID_PROFILES:
            fail(f"{name}: Profil '{prof}' ungueltig (erlaubt: {', '.join(VALID_PROFILES)}).")
        prof_l = prof
        _si = r.get("site_index")
        if _si is None or str(_si).strip()=="":
            if prof_l=="sirene":
                idx=sir_idx; sir_idx+=1
            else:
                idx=fzg_idx; fzg_idx+=1
        else:
            try:
                idx = int(r["site_index"])
            except (TypeError, ValueError):
                fail(f"{name}: site_index ist keine Zahl.")
        if not (1 <= idx <= 254):
            fail(f"{name}: site_index {idx} ausserhalb 1..254.")
        _ls = r.get("lan_subnet")
        sub = "" if _ls is None else str(_ls).strip()
        if not sub:
            if prof=="sirene":
                sub=str(sir_slots[sir_n]); sir_n+=1
            else:
                sub=str(fzg_slots[fzg_n]); fzg_n+=1
        try:
            net = ipaddress.ip_network(sub, strict=True)
        except ValueError as e:
            fail(f"{name}: lan_subnet '{sub}' ungueltig ({e}).")
        if name in names:
            fail(f"Doppelter Standortname: {name}")
        if idx in indexes:
            fail(f"Doppelter site_index: {idx} (bei {name})")
        for prev_name, prev in nets:
            if net.overlaps(prev):
                fail(f"Subnetz-Ueberlappung: {name} ({net}) mit {prev_name} ({prev})")
        names.add(name); indexes.add(idx); nets.append((name, net))
        sites.append({"name": name, "profile": prof,
                      "site_index": idx, "lan_subnet": sub})

    if not sites:
        fail("Keine Standorte in der Tabelle gefunden.")

    # --- Anker ---
    wa = wb["Anker"]
    arows = read_rows(wa, ["id", "path_id", "aktiv", "endpoint_host",
                           "endpoint_port", "hub_public_key", "metric"])
    anchors = []
    for r in arows:
        aid = str(r.get("id", "")).strip()
        if not aid:
            continue
        anchors.append({
            "id": aid, "path_id": int(r["path_id"]),
            "active": str(r.get("aktiv", "")).strip().lower() in ("ja", "yes", "true", "1"),
            "endpoint_host": str(r.get("endpoint_host", "")).strip(),
            "endpoint_port": int(r["endpoint_port"]),
            "hub_public_key": str(r.get("hub_public_key", "")).strip(),
            "metric": int(r["metric"]),
            "description": str(r.get("beschreibung", "") or "").strip(),
        })

    # --- Zentrale Netze ---
    routes = []
    if "Zentrale_Netze" in wb.sheetnames:
        for row in wb["Zentrale_Netze"].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                v = str(row[0]).strip()
                try:
                    ipaddress.ip_network(v, strict=True)
                    routes.append(v)
                except ValueError:
                    fail(f"Zentrale_Netze: '{v}' ist kein gueltiges Netz.")
    if routes:
        base["global"]["routes_central"] = routes

    # --- zusammenbauen + schreiben ---
    base["anchors"] = anchors
    base["sites"] = sites
    yaml.safe_dump(base, open(YAML, "w", encoding="utf-8"),
                   allow_unicode=True, sort_keys=False)
    print(f"OK: {len(sites)} Standorte, {len(anchors)} Anker, "
          f"{len(routes) or len(base['global']['routes_central'])} zentrale Netze.")
    print(f"Validierung bestanden -> {YAML} geschrieben.")
    print("Naechster Schritt: python3 generate.py")


if __name__ == "__main__":
    main()
