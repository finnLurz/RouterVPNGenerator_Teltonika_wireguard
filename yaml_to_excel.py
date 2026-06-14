#!/usr/bin/env python3
"""Schreibt die berechneten Werte aus standorte.yaml in standorte.xlsx zurueck,
damit die Tabelle den generierten Zustand anzeigt (nur Anzeige, keine Pflege)."""
import os, yaml
from openpyxl import load_workbook
HERE=os.path.dirname(os.path.abspath(__file__))
c=yaml.safe_load(open(os.path.join(HERE,"standorte.yaml")))
wb=load_workbook(os.path.join(HERE,"standorte.xlsx"))
# Anker-Pubkeys ins Blatt Anker
wa=wb["Anker"]; amap={a["id"]:a for a in c["anchors"]}
for row in wa.iter_rows(min_row=2):
    aid=row[0].value
    if aid in amap:
        row[5].value=amap[aid]["hub_public_key"]   # Spalte F
# Standort-Index/Subnetz ins Blatt Standorte (nach name gematcht)
ws=wb["Standorte"]; smap={s["name"]:s for s in c["sites"]}
for row in ws.iter_rows(min_row=3):
    nm=row[0].value
    if nm in smap:
        row[2].value=smap[nm]["site_index"]   # C
        row[3].value=smap[nm]["lan_subnet"]   # D
wb.save(os.path.join(HERE,"standorte.xlsx"))
print("standorte.xlsx aktualisiert (Anzeige).")
